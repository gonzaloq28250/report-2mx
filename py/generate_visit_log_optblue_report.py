#!/usr/bin/env python3
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config


def sanitize_field_name(field_name):
    if not field_name:
        return None
    field = field_name.strip()
    field = field.replace(' ', '_').replace('-', '_')
    field = field.lower()
    return field


def read_template_info(template_path):
    wb = load_workbook(template_path, read_only=True)

    if 'Match' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Match'")

    ws_match = wb['Match']
    mapping = {}
    for row in range(2, ws_match.max_row + 1):
        report_col = ws_match.cell(row, 1).value
        db_field = ws_match.cell(row, 3).value
        if report_col and db_field:
            mapping[str(report_col).strip()] = str(db_field).strip()

    if 'Template' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Template'")

    ws = wb['Template']

    header_row = None
    for row in range(1, 20):
        val = ws.cell(row, 1).value
        if val and 'Visitas' in str(val):
            header_row = row
            break
    if not header_row:
        header_row = 1

    headers = []
    for col in range(1, 200):
        val = ws.cell(header_row, col).value
        if val:
            headers.append(str(val).strip())
        else:
            break

    wb.close()
    return mapping, headers, header_row


def get_table_columns(cursor, table_name):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return {col[0].lower(): col[0] for col in cursor.fetchall()}


def get_data_from_db(db_fields, table_name):
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    valid_fields = [f for f in db_fields if f]
    if valid_fields:
        fields_str = ', '.join([f'`{f}`' for f in valid_fields])
    else:
        fields_str = '*'
    query = f"SELECT {fields_str} FROM `{table_name}` ORDER BY id DESC"
    cursor.execute(query)
    rows = cursor.fetchall()
    field_names = [d[0] for d in cursor.description]
    cursor.close()
    conn.close()
    return rows, field_names


def resolve_db_field(report_col, mapping, existing_columns):
    db_field_raw = mapping.get(report_col)
    if not db_field_raw:
        return None

    import re
    sanitized = sanitize_field_name(db_field_raw)
    if not sanitized:
        return None

    col_lower = {k: v for k, v in existing_columns.items()}

    found = col_lower.get(sanitized)
    if not found:
        for lk, lv in col_lower.items():
            if sanitized == lk.replace(' ', '_').replace('-', '_'):
                found = lv
                break
    if not found:
        base = re.sub(r'_\d+$', '', sanitized)
        for lk, lv in col_lower.items():
            if base == re.sub(r'_\d+$', '', lk.replace(' ', '_').replace('-', '_')):
                found = lv
                break
    if not found:
        clean_db = sanitized.replace('_', '')
        for lk, lv in col_lower.items():
            clean_col = lk.replace('_', '').replace(' ', '')
            if clean_db == clean_col or (len(clean_db) > 3 and clean_db in clean_col):
                found = lv
                break

    if not found:
        return db_field_raw
    return found


def generate_excel():
    table_name = 'visit_log_optblue_report'
    template_path = config.TEMPLATES['visit_log_optblue']
    output_path = config.OUTPUT_FILES['visit_log_optblue']

    print(f"Template: {template_path}")

    match_mapping, template_headers, header_row = read_template_info(template_path)
    print(f"Mapeo Match: {len(match_mapping)} columnas")

    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    existing_columns = get_table_columns(cursor, table_name)
    cursor.close()
    conn.close()
    print(f"Columnas en tabla: {len(existing_columns)}")

    col_map = {}
    for hdr in template_headers:
        if hdr in match_mapping:
            resolved = resolve_db_field(hdr, match_mapping, existing_columns)
            if resolved:
                col_map[hdr] = resolved

    db_fields = list(col_map.values())
    rows, field_names = get_data_from_db(db_fields, table_name)
    print(f"Filas obtenidas: {len(rows)}")

    field_index = {name: idx for idx, name in enumerate(field_names)}

    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)

    sheet_title = 'Template'
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df = pd.DataFrame(columns=template_headers)
        df.to_excel(writer, sheet_name=sheet_title, index=False, startrow=header_row - 1)

        workbook = writer.book
        worksheet = writer.sheets[sheet_title]

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter'
        })
        data_fmt = workbook.add_format({'valign': 'vcenter'})

        for col_num, hdr in enumerate(template_headers):
            worksheet.write(header_row - 1, col_num, hdr, header_fmt)

        data_row = header_row
        for row_data in rows:
            for col_num, hdr in enumerate(template_headers):
                if hdr in col_map:
                    db_field = col_map[hdr]
                    idx = field_index.get(db_field)
                    value = row_data[idx] if idx is not None and idx < len(row_data) else ''
                else:
                    value = ''

                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    worksheet.write_datetime(data_row, col_num, value, data_fmt)
                    continue
                elif isinstance(value, (int, float)):
                    pass
                else:
                    value = str(value) if value else ''

                worksheet.write(data_row, col_num, value, data_fmt)
            data_row += 1

    print(f"Excel generado: {output_path}")
    return output_path


if __name__ == '__main__':
    generate_excel()
