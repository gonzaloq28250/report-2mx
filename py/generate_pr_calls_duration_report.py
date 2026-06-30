#!/usr/bin/env python3
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['pr_calls_duration']
OUTPUT_PATH = config.OUTPUT_FILES['pr_calls_duration']


def sanitize_field_name(field_name):
    if not field_name:
        return None
    import re
    field = field_name.strip().lower()
    field = field.replace(' ', '_').replace('-', '_').replace('/', '_')
    field = re.sub(r'[^\w]', '', field)
    return field


def strip_numeric_suffix(name):
    import re
    return re.sub(r'_\d+$', '', name)


def read_template_info(template_path):
    wb = load_workbook(template_path, read_only=True)

    if 'Match' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Match'")

    ws_match = wb['Match']
    mapping = {}
    fixed_values = {}
    for row in range(2, ws_match.max_row + 1):
        report_col = ws_match.cell(row, 1).value
        db_field = ws_match.cell(row, 3).value
        fixed_val = ws_match.cell(row, 4).value
        if not report_col:
            break
        report_col = str(report_col).strip()
        if isinstance(fixed_val, str) and fixed_val.strip():
            fixed_values[report_col] = fixed_val.strip()
            continue
        if db_field:
            mapping[report_col] = str(db_field).strip()

    target_sheet = None
    for name in wb.sheetnames:
        if name != 'Match':
            target_sheet = name
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]
    headers = []
    for col in range(1, 200):
        val = ws.cell(1, col).value
        if val:
            headers.append(str(val).strip())
        else:
            break

    wb.close()
    return mapping, fixed_values, headers, target_sheet


def get_table_columns(cursor, table_name):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return {col[0].lower(): col[0] for col in cursor.fetchall()}


def get_data_from_db(db_fields, table_name):
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    valid_fields = [f for f in db_fields if f]
    if valid_fields:
        fields_str = ', '.join([f'`{f}`' for f in valid_fields])
    else:
        fields_str = '*'
    query = f"SELECT {fields_str} FROM `{table_name}` ORDER BY id DESC"
    cursor.execute(query)
    rows = cursor.fetchall()
    field_names = [d[0] for d in cursor.description]
    cursor.close()
    conn.close()
    return rows, field_names


def resolve_db_field(report_col, mapping, existing_columns):
    db_field_raw = mapping.get(report_col)
    if not db_field_raw:
        return None

    sanitized = sanitize_field_name(db_field_raw)
    col_lower = {k: v for k, v in existing_columns.items()}

    found = col_lower.get(sanitized)
    if not found:
        for lk, lv in col_lower.items():
            if sanitized == lk.replace(' ', '_').replace('-', '_'):
                found = lv
                break
    if not found:
        base = strip_numeric_suffix(sanitized)
        for lk, lv in col_lower.items():
            if base == strip_numeric_suffix(lk.replace(' ', '_').replace('-', '_')):
                found = lv
                break
    if not found:
        clean_db = sanitized.replace('_', '')
        for lk, lv in col_lower.items():
            clean_col = lk.replace('_', '').replace(' ', '')
            if clean_db == clean_col or (len(clean_db) > 3 and clean_db in clean_col):
                found = lv
                break

    if not found:
        return None
    return found


def generate_excel():
    table_name = 'pr_calls_duration_report'
    print(f"Template: {TEMPLATE_PATH}")

    match_mapping, fixed_values, template_headers, target_sheet = read_template_info(TEMPLATE_PATH)
    print(f"Mapeo Match: {len(match_mapping)} columnas, valores fijos: {len(fixed_values)}")

    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    existing_columns = get_table_columns(cursor, table_name)
    cursor.close()
    conn.close()
    print(f"Columnas en tabla: {len(existing_columns)}")

    col_map = {}
    for hdr in template_headers:
        if hdr in match_mapping:
            resolved = resolve_db_field(hdr, match_mapping, existing_columns)
            if resolved:
                col_map[hdr] = resolved

    db_fields = list(col_map.values())
    rows, field_names = get_data_from_db(db_fields, table_name)
    print(f"Total registros: {len(rows)}")

    field_index = {name: idx for idx, name in enumerate(field_names)}

    output_dir = os.path.dirname(OUTPUT_PATH)
    os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_PATH, engine='xlsxwriter') as writer:
        df = pd.DataFrame(columns=template_headers)
        df.to_excel(writer, sheet_name=target_sheet, index=False)

        workbook = writer.book
        worksheet = writer.sheets[target_sheet]

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter'
        })
        data_fmt = workbook.add_format({'valign': 'vcenter'})

        for col_num, hdr in enumerate(template_headers):
            worksheet.write(0, col_num, hdr, header_fmt)

        data_row = 1
        for row_data in rows:
            for col_num, hdr in enumerate(template_headers):
                if hdr in fixed_values:
                    value = fixed_values[hdr]
                elif hdr in col_map:
                    db_field = col_map[hdr]
                    idx = field_index.get(db_field)
                    value = row_data[idx] if idx is not None and idx < len(row_data) else ''
                else:
                    value = ''

                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    worksheet.write_datetime(data_row, col_num, value, data_fmt)
                    continue
                elif isinstance(value, (int, float)):
                    pass
                else:
                    value = str(value) if value else ''

                worksheet.write(data_row, col_num, value, data_fmt)
            data_row += 1

    print(f"Excel generado: {OUTPUT_PATH}")
    return OUTPUT_PATH, len(rows)


if __name__ == '__main__':
    try:
        output_path, count = generate_excel()
        print(f"\nExcel generado exitosamente!")
        print(f"Archivo: {output_path}")
        print(f"Total registros: {count}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
