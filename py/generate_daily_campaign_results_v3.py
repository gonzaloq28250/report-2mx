"""
Generar Daily_Campaign_Results.xlsx desde MySQL
VERSION 3 - Nueva estructura de support_tables (9 columnas)
- Usa template v2: Daily-Campaign-Results-v2.xlsx
- Nueva estructura: disposition, final_disposition, redial, consultation,
                    agent_completes, presentations, unworkable, no_contact, net_contacts
- Queries actualizados según SUMMARY_ANALYSIS_V4.md
"""
import mysql.connector
from pathlib import Path
from datetime import datetime
import pandas as pd
from xlsx_utils import safe_save, load_template
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from copy import copy
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['daily_campaign']
OUTPUT_DIR = config.REPORTS_DIR
OUTPUT_DIR.mkdir(exist_ok=True)


def get_connection():
    """Crear conexion a MySQL"""
    return mysql.connector.connect(**config.MYSQL_CONFIG)


def get_list_name_records_map(conn):
    """
    Obtener mapas de list_name -> list_records desde list_name_records
    Si esta vacia, migra desde list_status_summary_report (tabla del watcher)
    Retorna dict {list_name: list_records}
    """
    cursor = conn.cursor()

    # Asegurar que list_name_records existe
    cursor.execute("CREATE TABLE IF NOT EXISTS list_name_records (id INT AUTO_INCREMENT PRIMARY KEY, list_name VARCHAR(255) NOT NULL, list_records INT NOT NULL DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP, UNIQUE KEY (list_name))")

    cursor.execute("SELECT COUNT(*) FROM list_name_records")
    count = cursor.fetchone()[0]

    if count == 0:
        cursor.execute("SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = DATABASE() AND table_name = 'list_status_summary_report'")
        if cursor.fetchone()[0] > 0:
            cursor.execute("SELECT COUNT(*) FROM list_status_summary_report")
            src_count = cursor.fetchone()[0]
            if src_count > 0:
                print(f"  list_name_records vacia, migrando {src_count} filas desde list_status_summary_report...")
                cursor.execute("""
                    SELECT lead_list_name, SUM(CAST(list_records AS UNSIGNED)) as total_records
                    FROM list_status_summary_report
                    WHERE lead_list_name IS NOT NULL AND lead_list_name != ''
                    GROUP BY lead_list_name
                """)
                for row in cursor.fetchall():
                    cursor.execute(
                        "INSERT INTO list_name_records (list_name, list_records) VALUES (%s, %s) "
                        "ON DUPLICATE KEY UPDATE list_records = VALUES(list_records)",
                        (row[0].strip(), row[1])
                    )
                conn.commit()
                cursor.execute("SELECT COUNT(*) FROM list_name_records")
                print(f"  Migradas {cursor.fetchone()[0]} filas a list_name_records")

    cursor.close()

    query = """
    SELECT list_name, list_records
    FROM list_name_records
    """
    df = pd.read_sql(query, conn)

    result = {}
    for _, row in df.iterrows():
        list_name = row['list_name'].strip()
        list_records = row['list_records']
        result[list_name] = list_records

    return result


def get_support_tables_map(conn):
    """
    Obttener mapas de dispositions desde support_tables nueva estructura (9 columnas)
    Retorna dict con listas de dispositions por cada categoría
    """
    query = """
    SELECT disposition, final_disposition, redial, consultation,
           agent_completes, presentations, unworkable, no_contact, net_contacts
    FROM support_tables
    ORDER BY disposition
    """
    df = pd.read_sql(query, conn)

    result = {
        'final_disposition': [],
        'redial': [],
        'consultation': [],
        'agent_completes': [],
        'presentations': [],
        'unworkable': [],
        'no_contact': [],
        'net_contacts': []
    }

    for _, row in df.iterrows():
        disp = row['disposition'].strip()
        if row['final_disposition']:
            result['final_disposition'].append(disp)
        if row['redial']:
            result['redial'].append(disp)
        if row['consultation']:
            result['consultation'].append(disp)
        if row['agent_completes']:
            result['agent_completes'].append(disp)
        if row['presentations']:
            result['presentations'].append(disp)
        if row['unworkable']:
            result['unworkable'].append(disp)
        if row['no_contact']:
            result['no_contact'].append(disp)
        if row['net_contacts']:
            result['net_contacts'].append(disp)

    # Log de carga
    print("  Support tables cargados:")
    for key, values in result.items():
        if values:
            print(f"    {key:20s}: {len(values)} dispositions")

    return result


def get_summary_data(conn, st_map):
    """
    Obtener datos para la hoja Summary
    Queries actualizados según SUMMARY_ANALYSIS_V4.md
    """
    # A - Date Range, C - Spanish Hours, D - English Hours (solo 2026)
    query_hours = """
    SELECT
        DATE(date) as fecha,
        SUM(CASE WHEN skill_availability LIKE '%PR%' THEN TIME_TO_SEC(login_time) ELSE 0 END) / 3600.0 as spanish_hours,
        SUM(CASE WHEN skill_availability LIKE '%VI%' THEN TIME_TO_SEC(login_time) ELSE 0 END) / 3600.0 as english_hours
    FROM login_logout
    WHERE date IS NOT NULL
      AND YEAR(date) = 2026
    GROUP BY DATE(date)
    ORDER BY fecha DESC
    """
    hours_df = pd.read_sql(query_hours, conn)

    # E - Total Dials (todos los records de outbound_call_log, solo 2026)
    query_dials = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as total_dials
    FROM outbound_call_log
    WHERE date IS NOT NULL
      AND YEAR(date) = 2026
    GROUP BY DATE(date)
    """
    dials_df = pd.read_sql(query_dials, conn)

    # G - Total Contacts (NOT IN no_contact)
    # Usar COLLATE para evitar problemas de collation entre tablas
    query_contacts = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as total_contacts
    FROM outbound_call_log o
    LEFT JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND (s.no_contact IS NULL OR s.no_contact = 0)
    GROUP BY DATE(date)
    """
    contacts_df = pd.read_sql(query_contacts, conn)

    # H - Net Contacts (usando support_tables.net_contacts)
    query_net_contacts = """
    SELECT
        DATE(o.date) as fecha,
        COUNT(*) as net_contacts
    FROM outbound_call_log o
    JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND s.net_contacts = 1
    GROUP BY DATE(o.date)
    """
    net_contacts_df = pd.read_sql(query_net_contacts, conn)

    # L - Total Presentations (WHERE presentations=1)
    query_presentations = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as total_presentations
    FROM outbound_call_log o
    JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND s.presentations = 1
    GROUP BY DATE(date)
    """
    presentations_df = pd.read_sql(query_presentations, conn)

    # P - Total Consultations (WHERE consultation=1)
    query_consultations = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as total_consultations
    FROM outbound_call_log o
    JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND s.consultation = 1
    GROUP BY DATE(date)
    """
    consultations_df = pd.read_sql(query_consultations, conn)

    # T/V/X - Agent Completes / Total Completes / Completes W/O EOC (WHERE agent_completes=1)
    query_completes = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as agent_completes
    FROM outbound_call_log o
    JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND s.agent_completes = 1
    GROUP BY DATE(date)
    """
    completes_df = pd.read_sql(query_completes, conn)

    # Z - Unworkables (WHERE unworkable=1)
    query_unworkables = """
    SELECT
        DATE(date) as fecha,
        COUNT(*) as unworkables
    FROM outbound_call_log o
    JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
      AND s.unworkable = 1
    GROUP BY DATE(date)
    """
    unworkables_df = pd.read_sql(query_unworkables, conn)

    # Combinar todos los datos
    result = []
    for _, row_hours in hours_df.iterrows():
        fecha = row_hours['fecha']
        spanish_hours = row_hours['spanish_hours'] or 0
        english_hours = row_hours['english_hours'] or 0
        total_hours = spanish_hours + english_hours

        # Buscar datos en otros dataframes
        dials_row = dials_df[dials_df['fecha'] == fecha]
        contacts_row = contacts_df[contacts_df['fecha'] == fecha]
        net_contacts_row = net_contacts_df[net_contacts_df['fecha'] == fecha]
        presentations_row = presentations_df[presentations_df['fecha'] == fecha]
        consultations_row = consultations_df[consultations_df['fecha'] == fecha]
        completes_row = completes_df[completes_df['fecha'] == fecha]
        unworkables_row = unworkables_df[unworkables_df['fecha'] == fecha]

        total_dials = dials_row.iloc[0]['total_dials'] if not dials_row.empty else 0
        total_contacts = contacts_row.iloc[0]['total_contacts'] if not contacts_row.empty else 0
        net_contacts = net_contacts_row.iloc[0]['net_contacts'] if not net_contacts_row.empty else 0
        total_presentations = presentations_row.iloc[0]['total_presentations'] if not presentations_row.empty else 0
        total_consultations = consultations_row.iloc[0]['total_consultations'] if not consultations_row.empty else 0
        agent_completes = completes_row.iloc[0]['agent_completes'] if not completes_row.empty else 0
        total_completes = agent_completes  # V = T
        completes_wo_eoc = agent_completes  # X = T
        unworkables = unworkables_row.iloc[0]['unworkables'] if not unworkables_row.empty else 0

        # Fórmulas calculadas
        dph = total_dials / total_hours if total_hours > 0 else 0
        cph = total_contacts / total_hours if total_hours > 0 else 0
        net_cph = net_contacts / total_hours if total_hours > 0 else 0
        contact_rate = net_contacts / total_dials if total_dials > 0 else 0
        pph = total_presentations / total_hours if total_hours > 0 else 0
        presentations_to_total = total_presentations / total_contacts if total_contacts > 0 else 0
        presentations_to_net = total_presentations / net_contacts if net_contacts > 0 else 0
        consultations_per_hour = total_consultations / total_hours if total_hours > 0 else 0
        consultations_to_total = total_consultations / total_contacts if total_contacts > 0 else 0
        consultations_to_net = total_consultations / net_contacts if net_contacts > 0 else 0
        agent_completes_per_hour = agent_completes / total_hours if total_hours > 0 else 0
        total_completes_per_hour = total_completes / total_hours if total_hours > 0 else 0
        completes_wo_eoc_per_hour = completes_wo_eoc / total_hours if total_hours > 0 else 0
        unworkable_rate = unworkables / completes_wo_eoc if completes_wo_eoc > 0 else 0

        record = {
            'fecha': fecha,
            'spanish_hours': spanish_hours,
            'english_hours': english_hours,
            'total_hours': total_hours,
            'total_dials': total_dials,
            'dph': dph,
            'total_contacts': total_contacts,
            'net_contacts': net_contacts,
            'cph': cph,
            'net_cph': net_cph,
            'contact_rate': contact_rate,
            'total_presentations': total_presentations,
            'pph': pph,
            'presentations_to_total': presentations_to_total,
            'presentations_to_net': presentations_to_net,
            'total_consultations': total_consultations,
            'consultations_per_hour': consultations_per_hour,
            'consultations_to_total': consultations_to_total,
            'consultations_to_net': consultations_to_net,
            'agent_completes': agent_completes,
            'agent_completes_per_hour': agent_completes_per_hour,
            'total_completes': total_completes,
            'total_completes_per_hour': total_completes_per_hour,
            'completes_wo_eoc': completes_wo_eoc,
            'completes_wo_eoc_per_hour': completes_wo_eoc_per_hour,
            'unworkables': unworkables,
            'unworkable_rate': unworkable_rate
        }
        result.append(record)

    return result


def get_campaign_data(conn, st_map):
    """
    Obtener datos por campana para cada fecha
    Queries actualizados con nueva estructura de support_tables
    """
    query = """
    SELECT
        DATE(o.date) as fecha,
        COALESCE(o.lead_list_name, 'Unknown') as list_name,
        COUNT(*) as total_dials,
        COUNT(CASE WHEN s.no_contact IS NULL OR s.no_contact = 0 THEN 1 END) as total_contacts,
        COUNT(CASE WHEN s.net_contacts = 1 THEN 1 END) as net_contacts,
        COUNT(CASE WHEN s.presentations = 1 THEN 1 END) as presentations,
        COUNT(CASE WHEN s.consultation = 1 THEN 1 END) as consultations,
        COUNT(CASE WHEN s.agent_completes = 1 THEN 1 END) as agent_completes,
        COUNT(CASE WHEN s.unworkable = 1 THEN 1 END) as unworkables
    FROM outbound_call_log o
    LEFT JOIN support_tables s ON o.disposition COLLATE utf8mb4_unicode_ci = s.disposition COLLATE utf8mb4_unicode_ci
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
    GROUP BY DATE(o.date), o.lead_list_name
    ORDER BY fecha DESC, list_name
    """

    df = pd.read_sql(query, conn)

    # Obtener Production Hours por fecha y skill (para prorratear)
    query_hours = """
    SELECT
        DATE(date) as fecha,
        SUM(CASE WHEN skill_availability LIKE '%PR%' THEN TIME_TO_SEC(login_time) ELSE 0 END) / 3600.0 as spanish_hours,
        SUM(CASE WHEN skill_availability LIKE '%VI%' THEN TIME_TO_SEC(login_time) ELSE 0 END) / 3600.0 as english_hours
    FROM login_logout
    WHERE date IS NOT NULL
    GROUP BY DATE(date)
    """
    hours_df = pd.read_sql(query_hours, conn)
    hours_map = {}
    for _, hr in hours_df.iterrows():
        fecha = hr['fecha']
        hours_map[fecha] = {
            'spanish': float(hr['spanish_hours'] or 0),
            'english': float(hr['english_hours'] or 0)
        }

    # Agregar horas a cada campana (prorrateado por dials)
    for idx, row in df.iterrows():
        fecha = row['fecha']
        total_spanish = hours_map.get(fecha, {}).get('spanish', 0)
        total_english = hours_map.get(fecha, {}).get('english', 0)
        total_dials_fecha = df[df['fecha'] == fecha]['total_dials'].sum()

        if total_dials_fecha > 0:
            dial_ratio = row['total_dials'] / total_dials_fecha
            df.loc[idx, 'spanish_hours'] = total_spanish * dial_ratio
            df.loc[idx, 'english_hours'] = total_english * dial_ratio
            df.loc[idx, 'total_hours'] = (total_spanish + total_english) * dial_ratio
        else:
            df.loc[idx, 'spanish_hours'] = 0
            df.loc[idx, 'english_hours'] = 0
            df.loc[idx, 'total_hours'] = 0

    return df


def fill_summary_sheet(ws, summary_data):
    """Llenar hoja Summary con datos diarios, mensuales, trimestrales y YTD - COMPLETAMENTE DINAMICO"""
    print("  Llenando Summary (dinamico)...")

    from datetime import datetime
    from collections import defaultdict

    # Buscar fila de header
    header_row = None
    for row in range(1, 10):
        if ws.cell(row, 1).value == 'Date Range':
            header_row = row
            break

    if not header_row:
        print("    No se encontro header en Summary")
        return

    # Limpiar datos existentes (filas debajo del header)
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(1, 28):
            cell = ws.cell(row, col)
            try:
                cell.value = None
            except AttributeError:
                pass

    # Ordenar datos por fecha ASC (mas antigua a mas nueva)
    summary_data_sorted = sorted(summary_data, key=lambda x: x['fecha'])

    # Escribir datos diarios (con fórmulas donde corresponda)
    data_row = header_row + 1
    for record in summary_data_sorted:
        # Datos estáticos (sin fórmulas)
        ws.cell(data_row, 1).value = record['fecha']  # A: Date Range
        ws.cell(data_row, 3).value = record['spanish_hours']  # C: Spanish Hours
        ws.cell(data_row, 4).value = record['english_hours']  # D: English Hours
        ws.cell(data_row, 5).value = record['total_dials']  # E: Total Dials
        ws.cell(data_row, 7).value = record['total_contacts']  # G: Total Contacts
        ws.cell(data_row, 8).value = record['net_contacts']  # H: Net Contacts
        ws.cell(data_row, 12).value = record['total_presentations']  # L: Total Presentations
        ws.cell(data_row, 16).value = record['total_consultations']  # P: Total Consultations
        ws.cell(data_row, 20).value = record['agent_completes']  # T: Agent Completes
        ws.cell(data_row, 22).value = record['total_completes']  # V: Total Completes
        ws.cell(data_row, 24).value = record['completes_wo_eoc']  # X: Completes W/O EOC
        ws.cell(data_row, 26).value = record['unworkables']  # Z: Unworkables

        # Fórmulas de Excel (se actualizan automáticamente)
        ws.cell(data_row, 2).value = f"=C{data_row}+D{data_row}"  # B: Total Hours = C + D
        ws.cell(data_row, 6).value = f"=E{data_row}/B{data_row}"  # F: DPH = E / B
        ws.cell(data_row, 9).value = f"=G{data_row}/B{data_row}"  # I: CPH = G / B
        ws.cell(data_row, 10).value = f"=H{data_row}/B{data_row}"  # J: Net CPH = H / B
        ws.cell(data_row, 11).value = f"=H{data_row}/E{data_row}"  # K: Contact Rate = H / E
        ws.cell(data_row, 13).value = f"=L{data_row}/B{data_row}"  # M: PPH = L / B
        ws.cell(data_row, 14).value = f"=L{data_row}/G{data_row}"  # N: Pres to Total = L / G
        ws.cell(data_row, 15).value = f"=L{data_row}/H{data_row}"  # O: Pres to Net = L / H
        ws.cell(data_row, 17).value = f"=P{data_row}/B{data_row}"  # Q: Consult/Hour = P / B
        ws.cell(data_row, 18).value = f"=P{data_row}/G{data_row}"  # R: Consult to Total = P / G
        ws.cell(data_row, 19).value = f"=P{data_row}/H{data_row}"  # S: Consult to Net = P / H
        ws.cell(data_row, 21).value = f"=T{data_row}/B{data_row}"  # U: Agent Comp/Hour = T / B
        ws.cell(data_row, 23).value = f"=V{data_row}/B{data_row}"  # W: Total Comp/Hour = V / B
        ws.cell(data_row, 25).value = f"=X{data_row}/B{data_row}"  # Y: Comp WO EOC/Hour = X / B
        ws.cell(data_row, 27).value = f"=Z{data_row}/X{data_row}"  # AA: Unworkable Rate = Z / X

        data_row += 1

    first_data_row = header_row + 1
    last_data_row = data_row - 1

    print(f"    Escritos {len(summary_data_sorted)} dias de datos (filas {first_data_row}-{last_data_row})")

    # Agrupar datos por mes y detectar year
    month_data = defaultdict(list)  # {month_name: [row_numbers]}
    year_data = defaultdict(set)    # {year: set of months}

    for idx, record in enumerate(summary_data_sorted):
        fecha = record['fecha']
        if isinstance(fecha, str):
            fecha = datetime.strptime(fecha, '%Y-%m-%d')
        month_name = fecha.strftime('%B')
        year = fecha.year
        row_num = first_data_row + idx
        month_data[month_name].append(row_num)
        year_data[year].add(month_name)

    # Obtener el year principal
    main_year = max(year_data.keys()) if year_data else datetime.now().year

    # Crear filas de resumen mensual (dinamico segun los meses con datos)
    current_row = last_data_row + 2  # Dejar una fila vacia

    month_order = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']

    monthly_rows = {}  # {month_name: row_number}

    for month in month_order:
        if month in month_data and month_data[month]:
            rows = month_data[month]
            start_row = min(rows)
            end_row = max(rows)

            # Escribir fila de mes
            ws.cell(current_row, 1).value = month
            ws.cell(current_row, 2).value = f"=SUM(C{current_row}:D{current_row})"
            ws.cell(current_row, 3).value = f"=SUM(C{start_row}:C{end_row})"
            ws.cell(current_row, 4).value = f"=SUM(D{start_row}:D{end_row})"
            ws.cell(current_row, 5).value = f"=SUM(E{start_row}:E{end_row})"
            ws.cell(current_row, 6).value = f"=E{current_row}/B{current_row}"
            ws.cell(current_row, 7).value = f"=SUM(G{start_row}:G{end_row})"
            ws.cell(current_row, 8).value = f"=SUM(H{start_row}:H{end_row})"
            ws.cell(current_row, 9).value = f"=G{current_row}/B{current_row}"
            ws.cell(current_row, 10).value = f"=H{current_row}/B{current_row}"
            ws.cell(current_row, 11).value = f"=H{current_row}/E{current_row}"
            ws.cell(current_row, 12).value = f"=SUM(L{start_row}:L{end_row})"
            ws.cell(current_row, 13).value = f"=L{current_row}/B{current_row}"
            ws.cell(current_row, 14).value = f"=L{current_row}/G{current_row}"
            ws.cell(current_row, 15).value = f"=L{current_row}/H{current_row}"
            ws.cell(current_row, 16).value = f"=SUM(P{start_row}:P{end_row})"
            ws.cell(current_row, 17).value = f"=P{current_row}/B{current_row}"
            ws.cell(current_row, 18).value = f"=P{current_row}/G{current_row}"
            ws.cell(current_row, 19).value = f"=P{current_row}/H{current_row}"
            ws.cell(current_row, 20).value = f"=SUM(T{start_row}:T{end_row})"
            ws.cell(current_row, 21).value = f"=T{current_row}/B{current_row}"
            ws.cell(current_row, 22).value = f"=SUM(V{start_row}:V{end_row})"
            ws.cell(current_row, 23).value = f"=V{current_row}/B{current_row}"
            ws.cell(current_row, 24).value = f"=SUM(X{start_row}:X{end_row})"
            ws.cell(current_row, 25).value = f"=X{current_row}/B{current_row}"
            ws.cell(current_row, 26).value = f"=SUM(Z{start_row}:Z{end_row})"
            ws.cell(current_row, 27).value = f"=Z{current_row}/X{current_row}"

            monthly_rows[month] = current_row
            current_row += 1

    # Definicion de trimestres
    quarters = {
        'Q1': ['January', 'February', 'March'],
        'Q2': ['April', 'May', 'June'],
        'Q3': ['July', 'August', 'September'],
        'Q4': ['October', 'November', 'December']
    }

    quarter_rows = {}  # {quarter_name: row_number}

    # Crear filas de quarters DINAMICAMENTE (solo si tienen todos los meses del quarter)
    for quarter, months in quarters.items():
        # Verificar si TODOS los meses del quarter tienen datos
        if all(month in monthly_rows for month in months):
            first_month_row = monthly_rows[months[0]]
            last_month_row = monthly_rows[months[-1]]

            ws.cell(current_row, 1).value = quarter
            ws.cell(current_row, 2).value = f"=SUM(C{current_row}:D{current_row})"
            ws.cell(current_row, 3).value = f"=SUM(C{first_month_row}:C{last_month_row})"
            ws.cell(current_row, 4).value = f"=SUM(D{first_month_row}:D{last_month_row})"
            ws.cell(current_row, 5).value = f"=SUM(E{first_month_row}:E{last_month_row})"
            ws.cell(current_row, 6).value = f"=E{current_row}/B{current_row}"
            ws.cell(current_row, 7).value = f"=SUM(G{first_month_row}:G{last_month_row})"
            ws.cell(current_row, 8).value = f"=SUM(H{first_month_row}:H{last_month_row})"
            ws.cell(current_row, 9).value = f"=G{current_row}/B{current_row}"
            ws.cell(current_row, 10).value = f"=H{current_row}/B{current_row}"
            ws.cell(current_row, 11).value = f"=H{current_row}/E{current_row}"
            ws.cell(current_row, 12).value = f"=SUM(L{first_month_row}:L{last_month_row})"
            ws.cell(current_row, 13).value = f"=L{current_row}/B{current_row}"
            ws.cell(current_row, 14).value = f"=L{current_row}/G{current_row}"
            ws.cell(current_row, 15).value = f"=L{current_row}/H{current_row}"
            ws.cell(current_row, 16).value = f"=SUM(P{first_month_row}:P{last_month_row})"
            ws.cell(current_row, 17).value = f"=P{current_row}/B{current_row}"
            ws.cell(current_row, 18).value = f"=P{current_row}/G{current_row}"
            ws.cell(current_row, 19).value = f"=P{current_row}/H{current_row}"
            ws.cell(current_row, 20).value = f"=SUM(T{first_month_row}:T{last_month_row})"
            ws.cell(current_row, 21).value = f"=T{current_row}/B{current_row}"
            ws.cell(current_row, 22).value = f"=SUM(V{first_month_row}:V{last_month_row})"
            ws.cell(current_row, 23).value = f"=V{current_row}/B{current_row}"
            ws.cell(current_row, 24).value = f"=SUM(X{first_month_row}:X{last_month_row})"
            ws.cell(current_row, 25).value = f"=X{current_row}/B{current_row}"
            ws.cell(current_row, 26).value = f"=SUM(Z{first_month_row}:Z{last_month_row})"
            ws.cell(current_row, 27).value = f"=Z{current_row}/X{current_row}"

            quarter_rows[quarter] = current_row
            current_row += 1

    # Fila vacia antes de YTD
    current_row += 1

    # YTD (suma de TODOS los meses disponibles del year)
    ytd_row = current_row

    if monthly_rows:
        first_month_row = list(monthly_rows.values())[0]
        last_month_row = list(monthly_rows.values())[-1]

        ws.cell(ytd_row, 1).value = f"{main_year} YTD"
        ws.cell(ytd_row, 2).value = f"=SUM(C{ytd_row},D{ytd_row})"
        ws.cell(ytd_row, 3).value = f"=SUM(C{first_month_row}:C{last_month_row})"
        ws.cell(ytd_row, 4).value = f"=SUM(D{first_month_row}:D{last_month_row})"
        ws.cell(ytd_row, 5).value = f"=SUM(E{first_month_row}:E{last_month_row})"
        ws.cell(ytd_row, 6).value = f"=E{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 7).value = f"=SUM(G{first_month_row}:G{last_month_row})"
        ws.cell(ytd_row, 8).value = f"=SUM(H{first_month_row}:H{last_month_row})"
        ws.cell(ytd_row, 9).value = f"=G{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 10).value = f"=H{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 11).value = f"=H{ytd_row}/E{ytd_row}"
        ws.cell(ytd_row, 12).value = f"=SUM(L{first_month_row}:L{last_month_row})"
        ws.cell(ytd_row, 13).value = f"=L{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 14).value = f"=L{ytd_row}/G{ytd_row}"
        ws.cell(ytd_row, 15).value = f"=L{ytd_row}/H{ytd_row}"
        ws.cell(ytd_row, 16).value = f"=SUM(P{first_month_row}:P{last_month_row})"
        ws.cell(ytd_row, 17).value = f"=P{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 18).value = f"=P{ytd_row}/G{ytd_row}"
        ws.cell(ytd_row, 19).value = f"=P{ytd_row}/H{ytd_row}"
        ws.cell(ytd_row, 20).value = f"=SUM(T{first_month_row}:T{last_month_row})"
        ws.cell(ytd_row, 21).value = f"=T{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 22).value = f"=SUM(V{first_month_row}:V{last_month_row})"
        ws.cell(ytd_row, 23).value = f"=V{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 24).value = f"=SUM(X{first_month_row}:X{last_month_row})"
        ws.cell(ytd_row, 25).value = f"=X{ytd_row}/B{ytd_row}"
        ws.cell(ytd_row, 26).value = f"=SUM(Z{first_month_row}:Z{last_month_row})"
        ws.cell(ytd_row, 27).value = f"=Z{ytd_row}/X{ytd_row}"

    # Filas adicionales YTD (metricas)
    current_row = ytd_row + 1
    ws.cell(current_row, 1).value = 'Penetration Rate - YTD'
    ws.cell(current_row, 2).value = f"=V{ytd_row}/(B{current_row+3}-Z{ytd_row})"
    current_row += 1

    ws.cell(current_row, 1).value = 'Net Contact to Callable Leads - YTD'
    ws.cell(current_row, 2).value = f"=H{ytd_row}/B{current_row+2}"
    current_row += 1

    ws.cell(current_row, 1).value = 'Total Callable Leads Received - YTD'
    ws.cell(current_row, 2).value = 21781  # TODO: Hacer dinamico desde DB
    current_row += 1

    ws.cell(current_row, 1).value = 'Total Workable Leads - YTD'
    ws.cell(current_row, 2).value = f"=B{current_row-1}-Z{ytd_row}"

    # Resumen de lo generado
    quarters_created = list(quarter_rows.keys())
    print(f"    Creadas secciones: {len(monthly_rows)} meses, quarters: {quarters_created}, YTD")


def create_campaign_sheet(wb, campaign_name, campaign_data):
    """
    Crear/actualizar una hoja para una campana
    Usa la estructura del template v2 (27 columnas A-AA)
    """
    print(f"  Procesando campana: {campaign_name}")

    # Verificar si la hoja ya existe
    if campaign_name in wb.sheetnames:
        ws = wb[campaign_name]
        # Encontrar fila de header
        header_row = None
        for row in range(1, 10):
            if ws.cell(row, 1).value == 'Date Range' or (ws.cell(row, 1).value):
                header_row = row
                break
        if not header_row:
            header_row = 2
        # Limpiar datos existentes (evitando MergedCells)
        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, 28):
                cell = ws.cell(row, col)
                try:
                    cell.value = None
                except AttributeError:
                    pass  # Skip MergedCells
        data_start_row = header_row + 1
    else:
        # Crear nueva hoja (copiando estructura del template)
        ws = wb.create_sheet(title=campaign_name[:31])

        # Copiar estructura de la hoja Summary (headers y formatos)
        summary_ws = wb['Summary']

        # Encontrar header row en Summary
        summary_header_row = None
        for row in range(1, 10):
            if summary_ws.cell(row, 1).value == 'Date Range':
                summary_header_row = row
                break

        # Fila 1: Nombre de campaña
        ws.cell(1, 1).value = campaign_name

        # Fila 2: Copiar headers y estilos desde Summary
        if summary_header_row:
            for col in range(1, 28):
                src_cell = summary_ws.cell(summary_header_row, col)
                dst_cell = ws.cell(2, col)

                # Copiar valor
                dst_cell.value = src_cell.value

                # Copiar estilos
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format

        data_start_row = 3

    # Escribir datos por fecha
    current_row = data_start_row

    if not campaign_data.empty:
        fechas_unicas = sorted(campaign_data['fecha'].unique())

        for fecha in fechas_unicas:
            fecha_data = campaign_data[campaign_data['fecha'] == fecha].iloc[0]

            total_hours = fecha_data.get('total_hours', 0) or 0
            spanish_hours = fecha_data.get('spanish_hours', 0) or 0
            english_hours = fecha_data.get('english_hours', 0) or 0
            total_dials = fecha_data.get('total_dials', 0) or 0
            total_contacts = fecha_data.get('total_contacts', 0) or 0
            net_contacts = fecha_data.get('net_contacts', 0) or 0
            presentations = fecha_data.get('presentations', 0) or 0
            consultations = fecha_data.get('consultations', 0) or 0
            agent_completes = fecha_data.get('agent_completes', 0) or 0
            unworkables = fecha_data.get('unworkables', 0) or 0

            # Escribir datos estáticos y fórmulas
            ws.cell(current_row, 1).value = fecha  # A: Date Range
            ws.cell(current_row, 3).value = spanish_hours  # C: Spanish Hours
            ws.cell(current_row, 4).value = english_hours  # D: English Hours
            ws.cell(current_row, 2).value = f"=C{current_row}+D{current_row}"  # B: Total Hours = C + D
            ws.cell(current_row, 5).value = total_dials  # E: Total Dials

            # Fórmulas de Excel
            ws.cell(current_row, 6).value = f"=E{current_row}/B{current_row}"  # F: DPH
            ws.cell(current_row, 7).value = total_contacts  # G: Total Contacts
            ws.cell(current_row, 8).value = net_contacts  # H: Net Contacts
            ws.cell(current_row, 9).value = f"=G{current_row}/B{current_row}"  # I: CPH
            ws.cell(current_row, 10).value = f"=H{current_row}/B{current_row}"  # J: Net CPH
            ws.cell(current_row, 11).value = f"=H{current_row}/E{current_row}"  # K: Contact Rate
            ws.cell(current_row, 12).value = presentations  # L: Presentations
            ws.cell(current_row, 13).value = f"=L{current_row}/B{current_row}"  # M: PPH
            ws.cell(current_row, 14).value = f"=L{current_row}/G{current_row}"  # N: Pres to Total
            ws.cell(current_row, 15).value = f"=L{current_row}/H{current_row}"  # O: Pres to Net
            ws.cell(current_row, 16).value = consultations  # P: Consultations
            ws.cell(current_row, 17).value = f"=P{current_row}/B{current_row}"  # Q: Consult/Hour
            ws.cell(current_row, 18).value = f"=P{current_row}/G{current_row}"  # R: Consult to Total
            ws.cell(current_row, 19).value = f"=P{current_row}/H{current_row}"  # S: Consult to Net
            ws.cell(current_row, 20).value = agent_completes  # T: Agent Completes
            ws.cell(current_row, 21).value = f"=T{current_row}/B{current_row}"  # U: Agent Comp/Hour
            ws.cell(current_row, 22).value = f"=T{current_row}"  # V: Total Completes = T
            ws.cell(current_row, 23).value = f"=V{current_row}/B{current_row}"  # W: Total Comp/Hour
            ws.cell(current_row, 24).value = f"=T{current_row}"  # X: Completes W/O EOC = T
            ws.cell(current_row, 25).value = f"=X{current_row}/B{current_row}"  # Y: Comp WO EOC/Hour
            ws.cell(current_row, 26).value = unworkables  # Z: Unworkables
            ws.cell(current_row, 27).value = f"=Z{current_row}/X{current_row}"  # AA: Unworkable Rate

            current_row += 1

    # Buscar o crear fila de Summary
    summary_row = None
    for row in range(current_row, current_row + 10):
        if ws.cell(row, 1).value == 'Summary':
            summary_row = row
            break

    if summary_row is None:
        summary_row = current_row  # Summary va inmediatamente después de los datos
        ws.cell(summary_row, 1).value = 'Summary'

    # Actualizar fila de Summary con fórmulas SUM
    separator_row = summary_row - 1
    ws.cell(summary_row, 3).value = f"=SUM(C{data_start_row}:C{separator_row})"  # Spanish Hours
    ws.cell(summary_row, 4).value = f"=SUM(D{data_start_row}:D{separator_row})"  # English Hours
    ws.cell(summary_row, 2).value = f"=C{summary_row}+D{summary_row}"  # Total Hours
    ws.cell(summary_row, 5).value = f"=SUM(E{data_start_row}:E{separator_row})"  # Total Dials
    ws.cell(summary_row, 6).value = f"=E{summary_row}/B{summary_row}"  # DPH
    ws.cell(summary_row, 7).value = f"=SUM(G{data_start_row}:G{separator_row})"  # Total Contacts
    ws.cell(summary_row, 8).value = f"=SUM(H{data_start_row}:H{separator_row})"  # Net Contacts
    ws.cell(summary_row, 9).value = f"=G{summary_row}/B{summary_row}"  # CPH
    ws.cell(summary_row, 10).value = f"=H{summary_row}/B{summary_row}"  # Net CPH
    ws.cell(summary_row, 11).value = f"=H{summary_row}/E{summary_row}"  # Contact Rate
    ws.cell(summary_row, 12).value = f"=SUM(L{data_start_row}:L{separator_row})"  # Presentations
    ws.cell(summary_row, 13).value = f"=L{summary_row}/B{summary_row}"  # PPH
    ws.cell(summary_row, 14).value = f"=L{summary_row}/G{summary_row}"  # Pres to Total
    ws.cell(summary_row, 15).value = f"=L{summary_row}/H{summary_row}"  # Pres to Net
    ws.cell(summary_row, 16).value = f"=SUM(P{data_start_row}:P{separator_row})"  # Consultations
    ws.cell(summary_row, 17).value = f"=P{summary_row}/B{summary_row}"  # Consult/Hour
    ws.cell(summary_row, 18).value = f"=P{summary_row}/G{summary_row}"  # Consult to Total
    ws.cell(summary_row, 19).value = f"=P{summary_row}/H{summary_row}"  # Consult to Net
    ws.cell(summary_row, 20).value = f"=SUM(T{data_start_row}:T{separator_row})"  # Agent Completes
    ws.cell(summary_row, 21).value = f"=T{summary_row}/B{summary_row}"  # Agent Comp/Hour
    ws.cell(summary_row, 22).value = f"=SUM(V{data_start_row}:V{separator_row})"  # Total Completes
    ws.cell(summary_row, 23).value = f"=V{summary_row}/B{summary_row}"  # Total Comp/Hour
    ws.cell(summary_row, 24).value = f"=SUM(X{data_start_row}:X{separator_row})"  # Completes W/O EOC
    ws.cell(summary_row, 25).value = f"=X{summary_row}/B{summary_row}"  # Comp WO EOC/Hour
    ws.cell(summary_row, 26).value = f"=SUM(Z{data_start_row}:Z{separator_row})"  # Unworkables
    ws.cell(summary_row, 27).value = f"=Z{summary_row}/X{summary_row}"  # AA: Unworkable Rate

    # Crear filas de metricas YTD despues de Summary
    ytd_metrics_row = summary_row + 2  # Dejar una fila vacia

    # Penetration Rate - YTD
    ws.cell(ytd_metrics_row, 1).value = 'Penetration Rate - YTD'
    ws.cell(ytd_metrics_row, 2).value = f"=V{summary_row}/(B{ytd_metrics_row+3}-Z{summary_row})"
    ytd_metrics_row += 1

    # Net Contact to Callable Leads - YTD
    ws.cell(ytd_metrics_row, 1).value = 'Net Contact to Callable Leads - YTD'
    ws.cell(ytd_metrics_row, 2).value = f"=H{summary_row}/B{ytd_metrics_row+2}"
    ytd_metrics_row += 1

    # Total Callable Leads Received - YTD (se actualiza con list_records despues)
    ws.cell(ytd_metrics_row, 1).value = 'Total Callable Leads Received - YTD'
    ws.cell(ytd_metrics_row, 2).value = 0  # Placeholder, se actualiza con list_records
    ytd_metrics_row += 1

    # Total Workable Leads - YTD
    ws.cell(ytd_metrics_row, 1).value = 'Total Workable Leads - YTD'
    ws.cell(ytd_metrics_row, 2).value = f"=B{ytd_metrics_row-1}-Z{summary_row}"

    return summary_row


def update_list_records_in_campaign_sheet(ws, campaign_name, list_name_records_map):
    """
    Actualizar Total Callable Leads Received - YTD con valor de list_name_records
    Busca la fila donde Columna A = "Total Callable Leads Received - YTD"
    y actualiza Columna B con el valor de list_records
    """
    print(f"    Actualizando list_records para {campaign_name}...")
    print(f"      Filas en sheet: {ws.max_row}")

    # Buscar la fila donde Columna A = "Total Callable Leads Received - YTD"
    for row in range(1, ws.max_row + 1):
        cell_a_value = ws.cell(row, 1).value
        if cell_a_value:
            cell_a_str = str(cell_a_value).strip()
            if 'Total Callable Leads Received - YTD' in cell_a_str:
                list_records = list_name_records_map.get(campaign_name)
                print(f"      Fila {row}: Encontrada! list_records = {list_records}")
                if list_records is not None:
                    old_value = ws.cell(row, 2).value
                    ws.cell(row, 2).value = list_records
                    print(f"      Actualizado: {old_value} -> {list_records}")
                break

    print(f"    Busqueda completada para {campaign_name}")


def update_campaign_summary(wb, campaign_summary_rows):
    """Actualizar Campaign Summary con descripciones dinámicas en H-L y datos de campañas en A-F

    Estructura:
    - Filas 2+: Cada fila tiene A-F (datos de campaña) + H-L (descripción del segmento)
    - Las descripciones en H-L ciclan a través de los segmentos que tienen datos
    """
    print("  Actualizando Campaign Summary...")

    if 'Campaign Summary' not in wb.sheetnames:
        print("    No existe hoja Campaign Summary")
        return

    ws = wb['Campaign Summary']

    # Limpiar datos existentes
    for row in range(2, 200):
        for col in range(1, 13):
            cell = ws.cell(row, col)
            try:
                cell.value = None
            except AttributeError:
                pass  # Skip MergedCells

    # Mapeo de segmentacion (orden: longest keys first to avoid partial matches)
    segmentation_map = {
        'TRENDING': 'Trending Inactive',
        'INACT': 'Inactive',
        'NYA': 'Never Active',
        'NS': 'New Signings',
        'ACT': 'Active'
    }

    # Obtener el orden real de las hojas de campaña
    campaign_sheet_order = [s for s in wb.sheetnames if s not in ['Summary', 'Campaign Summary']]

    # Agrupar campañas por segmentacion
    seg_campaigns = {
        'Active': [],
        'Inactive': [],
        'New Signings': [],
        'Never Active': [],
        'Trending Inactive': []
    }

    for campaign in campaign_sheet_order:
        if campaign not in campaign_summary_rows:
            continue

        summary_row = campaign_summary_rows[campaign]

        # Determinar segmentacion
        seg = None
        for key, value in segmentation_map.items():
            if key in campaign:
                seg = value
                break

        if seg:
            seg_campaigns[seg].append((campaign, summary_row))

    # Segmentos que tienen campañas (orden específico)
    segmentation_order = ['Active', 'Inactive', 'Never Active', 'New Signings', 'Trending Inactive']
    active_segments = [s for s in segmentation_order if seg_campaigns.get(s)]

    # Rastrear filas del Campaign Summary por segmento (para columnas D y E)
    seg_rows_in_summary = {seg: [] for seg in active_segments}

    # ESCRIBIR FILAS: Dos tablas independientes
    # Tabla 1: Columnas A-F (filas 2+) - Campañas individuales
    # Tabla 2: Columnas H-L (filas 2-5) - Segmentos (rollup)

    # Parte 1: Recopilar info de campañas y calcular fórmulas de segmento
    data_row = 2
    campaign_rows_info = []

    for campaign in campaign_sheet_order:
        if campaign not in campaign_summary_rows:
            continue

        summary_row = campaign_summary_rows[campaign]
        if summary_row is None:
            continue

        # Determinar segmentacion
        seg = 'Unknown'
        for key, value in segmentation_map.items():
            if key in campaign:
                seg = value
                break

        # Filas YTD
        penetration_row = summary_row + 2
        net_contact_row = summary_row + 3
        callable_row = summary_row + 4
        workable_row = summary_row + 5

        # Guardar info
        campaign_rows_info.append({
            'campaign': campaign,
            'seg': seg,
            'summary_row': summary_row,
            'penetration_row': penetration_row,
            'net_contact_row': net_contact_row,
            'callable_row': callable_row,
            'workable_row': workable_row,
            'data_row': data_row
        })

        # Guardar fila para cálculos de segmento (columnas D y E)
        if seg in seg_rows_in_summary:
            seg_rows_in_summary[seg].append(data_row)

        data_row += 1

    # Parte 2: Calcular fórmulas de rollup por segmento
    seg_rollup_formulas = {}
    for seg_name in active_segments:
        rows = seg_rows_in_summary.get(seg_name, [])

        if rows:
            col_d_range = ",".join([f"D{r}" for r in rows])
            col_e_range = ",".join([f"E{r}" for r in rows])

            sum_completes_parts = []
            sum_net_contacts_parts = []

            for campaign_info in campaign_rows_info:
                if campaign_info['seg'] == seg_name:
                    campaign_name = campaign_info['campaign']
                    summary_row = campaign_info['summary_row']
                    if summary_row is not None:
                        sum_completes_parts.append(f"'{campaign_name}'!V{summary_row}")
                        sum_net_contacts_parts.append(f"'{campaign_name}'!H{summary_row}")

            if sum_completes_parts:
                sum_completes = ",".join(sum_completes_parts)
                sum_net_contacts = ",".join(sum_net_contacts_parts)

                seg_rollup_formulas[seg_name] = {
                    'I': f"=SUM({sum_completes})/SUM({col_e_range})",
                    'J': f"=SUM({sum_net_contacts})/SUM({col_e_range})",
                    'K': f"=SUM({col_d_range})",
                    'L': f"=SUM({col_e_range})"
                }

    # Parte 3: Escribir TABLA 1 - Campañas (columnas A-F desde fila 2)
    for info in campaign_rows_info:
        data_row = info['data_row']

        # Copiar formato desde la fila 2 del template
        for col in range(1, 7):
            src_cell = ws.cell(2, col)
            dst_cell = ws.cell(data_row, col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

        # Columnas A-F: Datos de campaña
        ws.cell(data_row, 1).value = info['campaign']
        ws.cell(data_row, 2).value = f"='{info['campaign']}'!B{info['penetration_row']}"
        ws.cell(data_row, 3).value = f"='{info['campaign']}'!B{info['net_contact_row']}"
        ws.cell(data_row, 4).value = f"='{info['campaign']}'!B{info['callable_row']}"
        ws.cell(data_row, 5).value = f"='{info['campaign']}'!B{info['workable_row']}"
        ws.cell(data_row, 6).value = info['seg']

    # Parte 4: Escribir TABLA 2 - Segmentos (columnas H-L filas 2-5)
    seg_row = 2
    for seg_name in active_segments:
        ws.cell(seg_row, 8).value = seg_name  # H: Target Segmentation Roll up

        if seg_name in seg_rollup_formulas:
            formulas = seg_rollup_formulas[seg_name]
            ws.cell(seg_row, 9).value = formulas['I']  # I: Penetration Rate
            ws.cell(seg_row, 9).number_format = numbers.FORMAT_PERCENTAGE_00
            ws.cell(seg_row, 10).value = formulas['J']  # J: Net Contact Rate
            ws.cell(seg_row, 10).number_format = numbers.FORMAT_PERCENTAGE_00
            ws.cell(seg_row, 11).value = formulas['K']  # K: Total Callable Leads
            ws.cell(seg_row, 12).value = formulas['L']  # L: Total Workable Leads

        seg_row += 1

    print(f"    Actualizadas {len(campaign_sheet_order)} campañas y {len(active_segments)} segmentos")


def generate_daily_campaign_results_v3():
    """Función principal para generar el reporte v3"""
    print("=== Generando Daily_Campaign_Results.xlsx (V3) ===\n")

    conn = None
    try:
        # Conectar a MySQL
        print("Conectando a MySQL...")
        conn = get_connection()
        print("Conexión exitosa!\n")

        # Cargar datos de soporte
        print("Cargando support_tables (nueva estructura 9 columnas)...")
        st_map = get_support_tables_map(conn)

        # Cargar datos
        print("\nCargando datos desde MySQL...")
        summary_data = get_summary_data(conn, st_map)
        print(f"  Summary: {len(summary_data)} días")

        campaign_df = get_campaign_data(conn, st_map)
        print(f"  Campañas: {campaign_df['list_name'].nunique()} campañas únicas\n")

        campaign_names = sorted(campaign_df['list_name'].unique())

        # Cargar list_name_records
        print("Cargando list_name_records...")
        list_name_records_map = get_list_name_records_map(conn)
        print(f"  Registros encontrados: {len(list_name_records_map)} list_names\n")

        # Cargar template v2
        print(f"Cargando template v2: {TEMPLATE_PATH}")
        wb = load_template(TEMPLATE_PATH)

        # Llenar Summary
        if 'Summary' in wb.sheetnames:
            fill_summary_sheet(wb['Summary'], summary_data)

        # Procesar hojas de campaña
        print("\nProcesando hojas de campaña...")
        campaign_summary_rows = {}
        for campaign in campaign_names:
            campaign_data = campaign_df[campaign_df['list_name'] == campaign]
            summary_row = create_campaign_sheet(wb, campaign, campaign_data)
            campaign_summary_rows[campaign] = summary_row

            # Actualizar Total Callable Leads Received - YTD con valor de list_name_records
            if campaign in wb.sheetnames:
                update_list_records_in_campaign_sheet(wb[campaign], campaign, list_name_records_map)

        # Actualizar Campaign Summary
        print("\nActualizando Campaign Summary...")
        update_campaign_summary(wb, campaign_summary_rows)

        # Guardar
        output_path = config.OUTPUT_FILES['daily_campaign']
        safe_save(wb, output_path)

        print(f"\n[OK] Excel generado: {output_path}")
        print(f"  Hojas incluidas: Summary, Campaign Summary, {len(campaign_names)} hojas de campana")

        return str(output_path)

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        if conn:
            conn.close()
            print("\nConexión cerrada")


if __name__ == '__main__':
    generate_daily_campaign_results_v3()
