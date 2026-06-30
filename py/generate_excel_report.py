"""
Generar Service-Level-Report.xlsx desde MySQL para ICC Amex
Lee formato del template, escribe datos con xlsxwriter (sin corrupcion)
"""

import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xlsxwriter.utility import xl_col_to_name
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['service_level']
OUTPUT_PATH = config.OUTPUT_FILES['service_level']


def get_connection():
    return mysql.connector.connect(**config.MYSQL_CONFIG)


def get_service_level_data(conn):
    cursor = conn.cursor()

    query = """
    SELECT
        DATE(o.date) as fecha,
        COUNT(*) as total_dials,
        AVG(CASE WHEN o.talk_time IS NOT NULL AND o.talk_time != '00:00:00'
            THEN TIME_TO_SEC(o.talk_time) ELSE NULL END) as avg_talk_time_seconds,
        AVG(CASE WHEN o.after_call_work_time IS NOT NULL AND o.after_call_work_time != '00:00:00'
            THEN TIME_TO_SEC(o.after_call_work_time) ELSE NULL END) as avg_followup_seconds,
        AVG(CASE WHEN o.handle_time IS NOT NULL AND o.handle_time != '00:00:00'
            THEN TIME_TO_SEC(o.handle_time) ELSE NULL END) as avg_handle_time_seconds,
        COALESCE(MAX(l.production_hours), 0) as production_hours
    FROM outbound_call_log o
    LEFT JOIN (
        SELECT DATE(date) as ldate, SUM(TIME_TO_SEC(login_time)) / 3600.0 as production_hours
        FROM login_logout
        WHERE date IS NOT NULL AND login_time IS NOT NULL
        GROUP BY DATE(date)
    ) l ON DATE(o.date) = l.ldate
    WHERE o.date IS NOT NULL
      AND YEAR(o.date) = 2026
    GROUP BY DATE(o.date)
    ORDER BY DATE(o.date)
    """

    cursor.execute(query)
    results = cursor.fetchall()
    print(f"Registros obtenidos: {len(results)}")
    cursor.close()

    data = []
    for row in results:
        def sec_to_time(seconds):
            if seconds is None or seconds == 0:
                return '00:00:00'
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            secs = int(seconds % 60)
            return f'{hours:02d}:{minutes:02d}:{secs:02d}'

        data.append({
            'Date': row[0],
            'Total Dials': row[1] or 0,
            'OB Average Talk Time': sec_to_time(row[2]),
            'OB Average Followup': sec_to_time(row[3]),
            'OB Total Average Handle Time': sec_to_time(row[4]),
            'Production Hours': round(row[5] or 0, 2)
        })

    return data


def read_template_format(template_path):
    wb = load_workbook(template_path)
    ws = wb['Service Level']

    header_row = 1
    for row in range(1, 20):
        if ws.cell(row, 1).value == 'Date':
            header_row = row
            break

    headers = []
    col_widths = {}
    for col in range(1, 20):
        cell = ws.cell(header_row, col)
        val = cell.value
        if val:
            headers.append(str(val))
            col_letter = get_column_letter(col)
            if ws.column_dimensions[col_letter].width:
                col_widths[col - 1] = ws.column_dimensions[col_letter].width

    wb.close()
    return headers, col_widths


def generate_excel(output_path=None):
    print("=== Generando Service-Level-Report.xlsx ===\n")

    if output_path is None:
        output_path = OUTPUT_PATH

    try:
        conn = get_connection()
        sl_data = get_service_level_data(conn)
        df = pd.DataFrame(sl_data)

        template_headers, col_widths = read_template_format(TEMPLATE_PATH)

        output_dir = os.path.dirname(output_path)
        os.makedirs(output_dir, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Service Level', index=False)

            workbook = writer.book
            worksheet = writer.sheets['Service Level']

            header_fmt = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': '#FFFFFF',
                'border': 1,
                'text_wrap': True,
                'valign': 'vcenter'
            })

            for col_num, col_name in enumerate(df.columns):
                worksheet.write(0, col_num, col_name, header_fmt)

            for col_idx, width in col_widths.items():
                worksheet.set_column(col_idx, col_idx, width)

            date_fmt = workbook.add_format({'num_format': 'yyyy-mm-dd'})
            for row_idx in range(len(df)):
                worksheet.write_datetime(row_idx + 1, 0, df.iloc[row_idx]['Date'].to_pydatetime(), date_fmt)

        print(f"\nExcel generado: {output_path}")
        return str(output_path)

    except Exception as e:
        print(f"Error: {e}")
        raise
    finally:
        if 'conn' in locals():
            conn.close()


if __name__ == "__main__":
    generate_excel()
