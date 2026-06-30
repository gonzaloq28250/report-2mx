#!/usr/bin/env python3
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config


def sanitize_field_name(field_name):
    if not field_name:
        return None
    import re
    field = field_name.strip().lower()
    field = field.replace(' ', '_').replace('-', '_')
    field = re.sub(r'[^\w]', '', field)
    return field


def strip_numeric_suffix(name):
    import re
    return re.sub(r'_\d+$', '', name)


CALCULATED_FIELDS = {
    'month(entry_datesubmitted)': '__month__',
    'year(entry_datesubmitted)': '__year__',
}


def read_template_info(template_path):
    wb = load_workbook(template_path, read_only=True)

    if 'Match' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Match'")

    ws_match = wb['Match']
    mapping = {}
    defaults = {}
    for row in range(2, ws_match.max_row + 1):
        report_col = ws_match.cell(row, 1).value
        db_field = ws_match.cell(row, 3).value
        default_val = ws_match.cell(row, 4).value
        if not report_col:
            break
        report_col = str(report_col).strip()
        if isinstance(default_val, str) and default_val.strip():
            defaults[report_col] = default_val.strip()
        if db_field:
            mapping[report_col] = str(db_field).strip()

    if 'Template' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Template'")

    ws = wb['Template']
    headers = []
    for col in range(1, 200):
        val = ws.cell(1, col).value
        if val:
            headers.append(str(val).strip())
        else:
            break

    wb.close()
    return mapping, defaults, headers


def get_table_columns(cursor, table_name):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return {col[0].lower(): col[0] for col in cursor.fetchall()}


def get_data_from_db(db_fields, table_name):
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()

    real_fields = [f for f in db_fields if f and not f.startswith('__')]
    if not real_fields:
        fields_str = '*'
    else:
        fields_str = ', '.join([f'`{f}`' for f in real_fields])

    query = f"""
        SELECT {fields_str},
               MONTH(entry_datesubmitted) as report_month,
               YEAR(entry_datesubmitted) as report_year
        FROM `{table_name}`
        ORDER BY id DESC
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    field_names = [d[0] for d in cursor.description]
    cursor.close()
    conn.close()

    return rows, field_names


def resolve_db_field(report_col, mapping, existing_columns):
    db_field_raw = mapping.get(report_col)
    if not db_field_raw:
        return None, None

    db_field_lower = db_field_raw.lower().replace(' ', '')
    if db_field_lower in CALCULATED_FIELDS:
        return CALCULATED_FIELDS[db_field_lower], None

    sanitized = sanitize_field_name(db_field_raw)
    col_lower = {k: v for k, v in existing_columns.items()}

    found = col_lower.get(sanitized)
    if not found:
        for lk, lv in col_lower.items():
            if sanitized == lk.replace(' ', '_').replace('-', '_'):
                found = lv
                break
    if not found:
        base = strip_numeric_suffix(sanitized)
        for lk, lv in col_lower.items():
            if base == strip_numeric_suffix(lk.replace(' ', '_').replace('-', '_')):
                found = lv
                break
    if not found:
        clean_db = sanitized.replace('_', '')
        for lk, lv in col_lower.items():
            clean_col = lk.replace('_', '').replace(' ', '')
            if clean_db == clean_col or (len(clean_db) > 3 and clean_db in clean_col):
                found = lv
                break

    return found, db_field_raw


def generate_excel():
    table_name = 'calls_consolidate_report'
    template_path = config.TEMPLATES['calls_consolidate']
    output_path = config.OUTPUT_FILES['calls_consolidate']

    print(f"Template: {template_path}")

    match_mapping, defaults, template_headers = read_template_info(template_path)
    print(f"Mapeo Match: {len(match_mapping)} columnas, defaults: {len(defaults)}")

    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    existing_columns = get_table_columns(cursor, table_name)
    cursor.close()
    conn.close()
    print(f"Columnas en tabla: {len(existing_columns)}")

    col_map = {}
    db_fields_raw = []
    for hdr in template_headers:
        if hdr in match_mapping:
            resolved, raw = resolve_db_field(hdr, match_mapping, existing_columns)
            if resolved:
                col_map[hdr] = resolved
                db_fields_raw.append(raw if raw else resolved)

    rows, field_names = get_data_from_db(db_fields_raw, table_name)
    print(f"Filas obtenidas: {len(rows)}")

    field_index = {name: idx for idx, name in enumerate(field_names)}

    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df = pd.DataFrame(columns=template_headers)
        df.to_excel(writer, sheet_name='Template', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Template']

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter'
        })
        data_fmt = workbook.add_format({'valign': 'vcenter'})

        for col_num, hdr in enumerate(template_headers):
            worksheet.write(0, col_num, hdr, header_fmt)

        data_row = 1
        for row_data in rows:
            for col_num, hdr in enumerate(template_headers):
                if hdr in col_map:
                    db_field = col_map[hdr]
                    if db_field == '__month__':
                        idx = field_index.get('report_month')
                        value = row_data[idx] if idx is not None else ''
                    elif db_field == '__year__':
                        idx = field_index.get('report_year')
                        value = row_data[idx] if idx is not None else ''
                    elif db_field in field_index:
                        idx = field_index[db_field]
                        value = row_data[idx] if idx < len(row_data) else ''
                    else:
                        value = ''
                elif hdr in defaults:
                    value = defaults[hdr]
                else:
                    value = ''

                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    worksheet.write_datetime(data_row, col_num, value, data_fmt)
                    continue
                elif isinstance(value, (int, float)):
                    pass
                else:
                    value = str(value) if value else ''

                worksheet.write(data_row, col_num, value, data_fmt)
            data_row += 1

    print(f"Excel generado: {output_path}")
    return output_path


if __name__ == '__main__':
    generate_excel()
