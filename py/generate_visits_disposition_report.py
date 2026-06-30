#!/usr/bin/env python3
import mysql.connector
from py.xlsx_utils import safe_save, load_template
from openpyxl.styles import Font, PatternFill, Border, Alignment
from pathlib import Path
from datetime import datetime
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['visits_disposition']
OUTPUT_PATH = config.OUTPUT_FILES['visits_disposition']


def sanitize_field_name(field_name):
    if not field_name:
        return ''
    import re
    field = str(field_name).strip().lower()
    field = field.replace(' ', '_').replace('-', '_').replace('/', '_').replace('\\', '_')
    field = re.sub(r'[^\w]', '', field)
    field = re.sub(r'_+', '_', field)
    return field.strip('_')


def strip_numeric_suffix(name):
    import re
    return re.sub(r'_\d+$', '', name)


def read_match_mapping(wb, existing_columns):
    if 'Match' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Match'")
    ws = wb['Match']
    mapping = {}
    col_lower = {k: v for k, v in existing_columns.items()}

    for row in range(2, ws.max_row + 1):
        report_col = ws.cell(row, 1).value
        db_field = ws.cell(row, 3).value
        default_val = ws.cell(row, 4).value

        if not report_col:
            break

        report_col = str(report_col).strip()

        if default_val and str(default_val).strip().lower() in ('leave blank', 'none', ''):
            continue

        if db_field:
            db_field_raw = str(db_field).strip()
            sanitized_db = sanitize_field_name(db_field_raw)
            if not sanitized_db:
                continue

            found = col_lower.get(sanitized_db)
            if not found:
                for lk, lv in col_lower.items():
                    if sanitized_db == lk.replace(' ', '_').replace('-', '_'):
                        found = lv
                        break
            if not found:
                base = strip_numeric_suffix(sanitized_db)
                for lk, lv in col_lower.items():
                    if base == strip_numeric_suffix(lk.replace(' ', '_').replace('-', '_')):
                        found = lv
                        break
            if not found:
                clean_db = sanitized_db.replace('_', '')
                for lk, lv in col_lower.items():
                    clean_col = lk.replace('_', '').replace(' ', '')
                    if clean_db == clean_col or (len(clean_db) > 3 and clean_db in clean_col):
                        found = lv
                        break
            if found:
                mapping[report_col] = found
    return mapping


def get_table_columns(cursor, table_name):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return {col[0].lower(): col[0] for col in cursor.fetchall()}


def get_data_from_db(db_fields, table_name):
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    valid_fields = [f for f in db_fields if f]
    if valid_fields:
        fields_str = ', '.join([f'`{f}`' for f in valid_fields])
    else:
        fields_str = '*'
    query = f"""
        SELECT {fields_str},
               MONTH(STR_TO_DATE(entry_datesubmitted, '%Y-%m-%d %H:%i:%s.%f')) as report_month,
               YEAR(STR_TO_DATE(entry_datesubmitted, '%Y-%m-%d %H:%i:%s.%f')) as report_year
        FROM `{table_name}`
        ORDER BY id DESC
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows, valid_fields


def generate_excel():
    table_name = 'visits_disposition_report'
    print(f"Template: {TEMPLATE_PATH}")
    wb = load_template(TEMPLATE_PATH)

    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    existing_columns = get_table_columns(cursor, table_name)
    cursor.close()
    conn.close()
    print(f"Columnas en tabla: {len(existing_columns)}")

    match_mapping = read_match_mapping(wb, existing_columns)
    print(f"Mapeo Match: {len(match_mapping)} columnas")
    db_fields = list(match_mapping.values())
    rows, valid_fields = get_data_from_db(db_fields, table_name)
    print(f"Filas obtenidas: {len(rows)}")

    if 'Template' not in wb.sheetnames:
        raise ValueError("El template no tiene un sheet llamado 'Template'")
    ws = wb['Template']
    template_headers = []
    for col in range(1, min(ws.max_column + 1, 200)):
        val = ws.cell(1, col).value
        if val:
            template_headers.append(str(val).strip())

    col_mapping = {}
    for col_idx, header in enumerate(template_headers, start=1):
        if header in match_mapping:
            col_mapping[match_mapping[header]] = col_idx

    max_row_to_check = max(ws.max_row, 1000)
    for row in range(max_row_to_check, 1, -1):
        if row > 1:
            ws.delete_rows(row, 1)

    data_row = 2
    for db_row in rows:
        for col_idx, header_name in enumerate(template_headers, start=1):
            mysql_field = match_mapping.get(header_name)
            if mysql_field and mysql_field in valid_fields:
                field_pos = valid_fields.index(mysql_field)
                if field_pos < len(db_row):
                    value = db_row[field_pos]
                    cell = ws.cell(data_row, col_idx)
                    if value is None:
                        cell.value = ''
                    elif isinstance(value, datetime):
                        cell.value = value
                    elif isinstance(value, (int, float)):
                        cell.value = value
                    else:
                        cell.value = str(value) if value else ''

        ws.cell(data_row, 9).value = 'PR'
        data_row += 1

    clean_font = Font(name='Calibri', size=11, color='FF000000', bold=False, italic=False)
    clean_fill = PatternFill(fill_type=None)
    clean_border = Border()
    clean_alignment = Alignment(horizontal='general', vertical='bottom', wrap_text=False)
    for row in range(2, data_row):
        for col in range(1, len(template_headers) + 1):
            cell = ws.cell(row, col)
            cell.font = clean_font
            cell.fill = clean_fill
            cell.border = clean_border
            cell.alignment = clean_alignment

    for sheet_name in wb.sheetnames:
        if sheet_name != 'Template':
            wb.remove(wb[sheet_name])

    safe_save(wb, OUTPUT_PATH)
    print(f"Excel generado: {OUTPUT_PATH}")
    return OUTPUT_PATH, len(rows)


if __name__ == '__main__':
    try:
        output_path, count = generate_excel()
        print(f"\nExcel generado exitosamente!")
        print(f"Archivo: {output_path}")
        print(f"Total registros: {count}")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
