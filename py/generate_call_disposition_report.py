#!/usr/bin/env python3
import mysql.connector
from py.xlsx_utils import safe_save, load_template
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.styles.numbers import is_date_format as _is_date_format
from datetime import datetime
import sys, os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['call_disposition']
OUTPUT_PATH = config.OUTPUT_FILES['call_disposition']
TARGET_SHEET = 'Call_Disposition_Report'


def get_table_columns(cursor, table_name):
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return {col[0].lower(): col[0] for col in cursor.fetchall()}


def get_db_data(cursor, db_fields, table_name):
    valid_fields = [f for f in db_fields if f]
    fields_str = ', '.join([f'`{f}`' for f in valid_fields]) if valid_fields else '*'
    has_entry_date = any('entry_datesubmitted' in f.lower() for f in valid_fields)

    if has_entry_date:
        query = f"""
            SELECT {fields_str},
                   MONTH(STR_TO_DATE(entry_datesubmitted, '%Y-%m-%d %H:%i:%s.%f')) as report_month,
                   YEAR(STR_TO_DATE(entry_datesubmitted, '%Y-%m-%d %H:%i:%s.%f')) as report_year
            FROM `{table_name}`
            WHERE entry_datesubmitted IS NOT NULL AND entry_datesubmitted != ''
            ORDER BY id DESC
        """
    else:
        query = f"SELECT {fields_str} FROM `{table_name}` ORDER BY id DESC"

    cursor.execute(query)
    rows = cursor.fetchall()
    field_names = [d[0] for d in cursor.description]
    return rows, field_names


def generate_excel():
    table_name = 'call_disposition_report'
    print(f"Template: {TEMPLATE_PATH}")
    wb = load_template(TEMPLATE_PATH)

    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"El template no tiene sheet '{TARGET_SHEET}'")
    if 'Match' not in wb.sheetnames:
        raise ValueError("El template no tiene sheet 'Match'")

    # --- Read Match sheet: report column → DB field ---
    ws_match = wb['Match']
    match = {}
    for row in range(2, ws_match.max_row + 1):
        report_col = ws_match.cell(row, 1).value
        db_field = ws_match.cell(row, 2).value
        if report_col and db_field:
            report_col = str(report_col).strip()
            db_field = str(db_field).strip()
            match[report_col] = db_field

    print(f"Match entries: {len(match)}")

    # --- Read Call_Disposition_Report header → build column position map ---
    ws = wb[TARGET_SHEET]
    header = {}
    for cell in ws[1]:
        if cell.value:
            header[str(cell.value).strip()] = cell.column

    print(f"Columnas en reporte: {len(header)}")

    # --- Build position mapping: column_number → DB field to query ---
    # Also track Month/Year from entry_datesubmitted
    col_to_db = {}
    needs_month = False
    needs_year = False
    entry_date_field = None

    for report_col_name, db_field in match.items():
        col_num = header.get(report_col_name)
        if col_num is None:
            continue

        if db_field.lower().startswith('month('):
            needs_month = True
            entry_date_field = db_field[db_field.find('(')+1:db_field.find(')')]
            col_to_db[col_num] = ('month', entry_date_field)
        elif db_field.lower().startswith('year('):
            needs_year = True
            entry_date_field = db_field[db_field.find('(')+1:db_field.find(')')]
            col_to_db[col_num] = ('year', entry_date_field)
        else:
            col_to_db[col_num] = ('field', db_field)

    # Collect unique DB fields to query
    db_fields_set = set()
    for entry_type, field_name in col_to_db.values():
        if entry_type == 'field' and field_name:
            db_fields_set.add(field_name)

    print(f"DB fields to query: {len(db_fields_set)}")

    # --- Query data ---
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()

    # Add entry_datesubmitted if needed for month/year
    if (needs_month or needs_year) and entry_date_field:
        db_fields_set.add(entry_date_field)

    # Verify fields exist in table
    existing_columns = get_table_columns(cursor, table_name)
    valid_fields = []
    for f in db_fields_set:
        if f.lower() in existing_columns:
            valid_fields.append(existing_columns[f.lower()])
        else:
            print(f"  WARNING: field '{f}' not found in table")

    rows, field_names = get_db_data(cursor, valid_fields, table_name)
    cursor.close()
    conn.close()
    print(f"Total registros: {len(rows)}")

    # --- Build index map: field_name → column index in query result ---
    field_index = {name: idx for idx, name in enumerate(field_names)}

    # --- Clear existing data (rows 2+) ---
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                ws.cell(row, col_num).value = None

    # --- Fill data ---
    data_row = 2
    for row_data in rows:
        for col_num, (entry_type, field_name) in col_to_db.items():
            if entry_type == 'field':
                idx = field_index.get(field_name)
                if idx is not None:
                    value = row_data[idx]
                else:
                    value = None
            elif entry_type == 'month':
                idx = field_index.get('report_month')
                value = row_data[idx] if idx is not None else None
            elif entry_type == 'year':
                idx = field_index.get('report_year')
                value = row_data[idx] if idx is not None else None
            else:
                value = None

            cell = ws.cell(data_row, col_num)
            if value is None:
                cell.value = None
            elif isinstance(value, datetime):
                cell.value = value
            elif isinstance(value, (int, float)):
                if _is_date_format(cell.number_format):
                    cell.value = str(value)
                else:
                    cell.value = value
            else:
                cell.value = str(value) if value else None
        data_row += 1

    # --- Clean formatting ---
    clean_font = Font()
    clean_fill = PatternFill(fill_type=None)
    clean_border = Border()
    clean_alignment = Alignment(horizontal='left', vertical='center')
    for row in range(2, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row, col_num)
            if cell.value is not None:
                cell.font = clean_font
                cell.fill = clean_fill
                cell.border = clean_border
                cell.alignment = clean_alignment

    # --- Remove Match sheet, keep only target ---
    sheets_to_remove = [s for s in wb.sheetnames if s != TARGET_SHEET]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])

    safe_save(wb, OUTPUT_PATH)
    print(f"Excel generado: {OUTPUT_PATH}")
    return OUTPUT_PATH, len(rows)


if __name__ == '__main__':
    try:
        output_path, count = generate_excel()
        print(f"\nExcel generado exitosamente!")
        print(f"Archivo: {output_path}")
        print(f"Total registros: {count}")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
