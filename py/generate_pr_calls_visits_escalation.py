#!/usr/bin/env python3
import mysql.connector
from xlsx_utils import safe_save, load_template
from openpyxl.styles import PatternFill, Border, Alignment, Font
from pathlib import Path
from datetime import datetime
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_PATH = config.TEMPLATES['pr_escalation']
OUTPUT_PATH = config.OUTPUT_FILES['pr_escalation']

REPORT_MAPPING = {
    'Date': 'date',
    'Agent Name': 'agent_name',
    'Program': 'program',
    'Contact Name': 'contact_name',
    'Phone Number': 'phone_number',
    'Business Name': 'business_name',
    'SE Number': 'se_number',
    'Complaint Reason': 'complaint_reason',
    'Issue Description': 'issue_description',
    'Callback Requested?': 'callback_requested',
    'Call Resolution Notes': 'call_resolution_notes',
    'Resolved?': 'resolved',
}


def get_table_columns():
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    cursor.execute("SHOW COLUMNS FROM pr_calls_visits_escalation")
    columns = {col[0]: col[1] for col in cursor.fetchall()}
    cursor.close()
    conn.close()
    return columns


def get_data_from_db(from_date=None, to_date=None):
    conn = mysql.connector.connect(**config.MYSQL_CONFIG)
    cursor = conn.cursor()
    db_fields = [REPORT_MAPPING[report_col] for report_col in REPORT_MAPPING]
    fields_str = ', '.join([f'`{f}`' for f in db_fields])
    query = f"SELECT {fields_str} FROM pr_calls_visits_escalation"
    params = []
    if from_date and to_date:
        query += " WHERE date >= %s AND date <= %s"
        params.extend([from_date, to_date])
    query += " ORDER BY date DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows, db_fields


def generate_excel(from_date=None, to_date=None):
    print(f"Template: {TEMPLATE_PATH}")
    wb = load_template(TEMPLATE_PATH)
    existing_columns = get_table_columns()
    print(f"Columnas en tabla: {len(existing_columns)}")

    rows, db_fields = get_data_from_db(from_date, to_date)
    print(f"Filas obtenidas: {len(rows)}")

    sheet_name = 'PR Calls and Visits Escalation'
    for sn in wb.sheetnames:
        if sn.startswith('PR Calls and Visits Escalation'):
            sheet_name = sn
            break

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"El template no tiene el sheet '{sheet_name}'. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]

    col_mapping = {}
    for col in range(1, 20):
        header_value = ws.cell(1, col).value
        if header_value:
            header_str = str(header_value).strip()
            if header_str in REPORT_MAPPING:
                db_field = REPORT_MAPPING[header_str]
                if db_field in existing_columns:
                    col_mapping[db_field] = col

    print(f"Columnas mapeadas: {len(col_mapping)}")

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, db_field in enumerate(db_fields):
            if db_field in col_mapping:
                excel_col = col_mapping[db_field]
                value = row_data[col_idx]
                cell = ws.cell(row_idx, excel_col)
                if isinstance(value, datetime):
                    cell.value = value
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value) if value else None

    clean_fill = PatternFill(fill_type=None)
    clean_border = Border()
    clean_alignment = Alignment(horizontal='left', vertical='center')
    clean_font = Font()
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell.font = clean_font
                cell.fill = clean_fill
                cell.border = clean_border
                cell.alignment = clean_alignment

    target_name = sheet_name
    for sn in list(wb.sheetnames):
        if sn != target_name:
            wb.remove(wb[sn])

    safe_save(wb, OUTPUT_PATH)
    print(f"Excel generado: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == '__main__':
    from_date = None
    to_date = None
    if '--from' in sys.argv and '--to' in sys.argv:
        from_idx = sys.argv.index('--from')
        to_idx = sys.argv.index('--to')
        from_date = sys.argv[from_idx + 1] if from_idx + 1 < len(sys.argv) else None
        to_date = sys.argv[to_idx + 1] if to_idx + 1 < len(sys.argv) else None
    generate_excel(from_date, to_date)
