"""
Generar reportes XLSX desde tablas MySQL usando templates
Lee Match sheet con openpyxl, escribe con xlsxwriter (sin corrupcion)
"""
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import csv
import re
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

TEMPLATE_DIR = config.TEMPLATE_DIR
OUTPUT_DIR = config.REPORTS_DIR
OUTPUT_DIR.mkdir(exist_ok=True)


def sanitize_column_name(col_name):
    if not col_name or col_name.strip() == '':
        return 'column'
    name = col_name.strip()
    for ch in [' – ', '-', ' ', '/', '\\', '(', ')', '"', "'", '?', '!', ',', '.', ':', ';', '|']:
        name = name.replace(ch, '_')
    name = name.replace('__', '_').strip('_')
    name = re.sub(r'[^\w]', '', name).lower()
    if not name:
        return 'column'
    if len(name) > 64:
        last_us = name.rfind('_', 0, 60)
        name = name[:last_us] if last_us > 0 else name[:60]
    return name


def sanitize_column_names_unique(headers):
    sanitized = {}
    seen = {}
    suffix_counter = {}
    for header in headers:
        sname = sanitize_column_name(header)
        if sname in seen:
            if sname not in suffix_counter:
                suffix_counter[sname] = 1
            else:
                suffix_counter[sname] += 1
            new_name = f"{sname[:60]}_{suffix_counter[sname]}"
        else:
            new_name = sname
        seen[new_name] = True
        sanitized[header] = new_name
    return sanitized


def get_csv_headers(csv_filename):
    csv_files = list(FTP_CSV_DIR.glob(f'{csv_filename}*.csv'))
    if not csv_files:
        return None
    with open(csv_files[0], 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        return next(reader)


def get_csv_data(csv_filename):
    csv_files = list(FTP_CSV_DIR.glob(f'{csv_filename}*.csv'))
    if not csv_files:
        return None, None
    with open(csv_files[0], 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = list(reader)
    return [{headers[i]: row[i] if i < len(row) else '' for i in range(headers)} for row in rows], headers


def get_connection():
    return mysql.connector.connect(**config.MYSQL_CONFIG)


def get_table_name_from_template(template_name):
    name = template_name.replace('_2026.xlsx', '').replace('_2026.XLSX', '')
    mapping = {
        'Visit_Log_Report(Optblue)': 'visit_log_optblue',
        'Visits_Disposition_Report': 'visits_disposition_report',
        'Call_Disposition_Report': 'call_disposition_report',
        'Calls_Consolidate_Report': 'calls_consolidate_report'
    }
    return mapping.get(name, name.lower())


def read_match_sheet(template_path):
    wb = load_workbook(template_path, read_only=True)
    if 'Match' not in wb.sheetnames:
        wb.close()
        return None
    ws = wb['Match']
    mapping = {}
    row = 2
    while True:
        report_col = ws.cell(row, 1).value
        table_idx = ws.cell(row, 2).value
        if not report_col:
            break
        try:
            table_idx = int(table_idx) if table_idx else 0
        except (ValueError, TypeError):
            table_idx = 0
        mapping[report_col] = table_idx
        row += 1
    wb.close()
    return mapping


def read_match_sheet_with_mysql_columns(template_path):
    wb = load_workbook(template_path, read_only=True)
    if 'Match' not in wb.sheetnames:
        wb.close()
        return None
    ws = wb['Match']
    mapping = {}
    row = 2
    while True:
        report_col = ws.cell(row, 1).value
        mysql_col_name = ws.cell(row, 3).value
        if not report_col:
            break
        if mysql_col_name and str(mysql_col_name).strip():
            mapping[report_col] = str(mysql_col_name).strip()
        row += 1
    wb.close()
    return mapping


def get_template_headers(template_path):
    wb = load_workbook(template_path, read_only=True)
    ws = wb['Template']
    headers = []
    col = 1
    while True:
        val = ws.cell(1, col).value
        if not val:
            break
        headers.append(val)
        col += 1
    wb.close()
    return headers


def build_column_mapping(match_mapping, csv_headers, table_columns):
    final_mapping = {}
    csv_sanitized = sanitize_column_names_unique(csv_headers)
    for report_col, csv_idx in match_mapping.items():
        if csv_idx < 1 or csv_idx > len(csv_headers):
            continue
        csv_col_name = csv_headers[csv_idx - 1]
        mysql_col_name = csv_sanitized.get(csv_col_name)
        if mysql_col_name:
            try:
                mysql_idx = table_columns.index(mysql_col_name)
                final_mapping[report_col] = mysql_idx
            except ValueError:
                pass
    return final_mapping


def generate_report(template_name, csv_filename, conn):
    print(f"\n--- Generando: {template_name} ---")
    template_path = TEMPLATE_DIR / template_name
    if not template_path.exists():
        print(f"  [ERROR] Template no encontrado: {template_path}")
        return None

    table_name = get_table_name_from_template(template_name)
    print(f"  Tabla: {table_name}")

    use_mysql_names = template_name in ('Visits_Disposition_Report_2026.xlsx', 'Visit_Log_Report(Optblue)_2026.xlsx')

    if use_mysql_names:
        match_mapping = read_match_sheet_with_mysql_columns(template_path)
    else:
        match_mapping = read_match_sheet(template_path)

    if not match_mapping:
        print(f"  [ERROR] No se encontro sheet Match")
        return None
    print(f"  Mapeo Match: {len(match_mapping)} columnas")

    cursor = conn.cursor()
    cursor.execute(f"DESCRIBE `{table_name}`")
    table_columns = [row[0] for row in cursor.fetchall()]

    if use_mysql_names:
        final_mapping = {}
        for report_col, mysql_col_name in match_mapping.items():
            sname = sanitize_column_name(mysql_col_name)
            found_col = None
            if mysql_col_name in table_columns:
                found_col = mysql_col_name
            elif sname in table_columns:
                found_col = sname
            else:
                clean_compare = sname.replace('_', '')
                for tc in table_columns:
                    if clean_compare == tc.replace('_', '').lower():
                        found_col = tc
                        break
            if found_col:
                final_mapping[report_col] = found_col
    else:
        csv_data, csv_headers = get_csv_data(csv_filename)
        if csv_headers:
            final_mapping = build_column_mapping(match_mapping, csv_headers, table_columns)
        else:
            final_mapping = {}

    template_headers = get_template_headers(template_path)

    data_columns = [col for col in table_columns if col not in ['created_at', 'updated_at']]
    data_columns_sql = ', '.join([f'`{col}`' for col in data_columns])
    cursor.execute(f"SELECT {data_columns_sql} FROM `{table_name}` ORDER BY id")
    rows = cursor.fetchall()
    field_names = [d[0] for d in cursor.description]
    field_index = {name: idx for idx, name in enumerate(field_names)}
    print(f"  Filas obtenidas: {len(rows)}")

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_filename = f"{template_name.replace('.xlsx', '')}-{timestamp}.xlsx"
    output_path = OUTPUT_DIR / output_filename

    with pd.ExcelWriter(str(output_path), engine='xlsxwriter') as writer:
        df = pd.DataFrame(columns=template_headers)
        df.to_excel(writer, sheet_name='Template', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Template']

        header_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF',
            'border': 1, 'text_wrap': True, 'valign': 'vcenter'
        })
        data_fmt = workbook.add_format({'valign': 'vcenter'})

        for col_num, hdr in enumerate(template_headers):
            worksheet.write(0, col_num, hdr, header_fmt)

        data_row = 1
        for row_data in rows:
            for col_num, hdr in enumerate(template_headers):
                mysql_mapping = final_mapping.get(hdr)
                if mysql_mapping is None:
                    continue

                if use_mysql_names:
                    mysql_col_name = mysql_mapping
                    if mysql_col_name in data_columns:
                        col_index = data_columns.index(mysql_col_name)
                        value = row_data[col_index] if col_index < len(row_data) else ''
                    else:
                        value = ''
                else:
                    mysql_col_index = mysql_mapping
                    value = row_data[mysql_col_index] if 0 < mysql_col_index <= len(row_data) else ''

                if value is None:
                    value = ''

                if isinstance(value, datetime):
                    worksheet.write_datetime(data_row, col_num, value, data_fmt)
                elif isinstance(value, (int, float)):
                    worksheet.write(data_row, col_num, value, data_fmt)
                else:
                    worksheet.write(data_row, col_num, str(value) if value else '', data_fmt)

            if template_name == 'Visits_Disposition_Report_2026.xlsx':
                worksheet.write(data_row, 8, 'PR', data_fmt)

            data_row += 1

    print(f"  [OK] Reporte generado: {output_path}")
    return str(output_path)


def main():
    print("=== Generar Reportes FTP desde Templates ===")
    conn = get_connection()
    templates = [
        ('Calls_Consolidate_Report_2026.xlsx', 'Calls_Consolidate_Report'),
        ('Call_Disposition_Report_2026.xlsx', 'Call_Disposition_Report'),
        ('Visit_Log_Report(Optblue)_2026.xlsx', 'Visit_Log_(Optblue)'),
        ('Visits_Disposition_Report_2026.xlsx', 'Visits_Disposition_Report')
    ]
    generated = []
    for template, csv_base in templates:
        try:
            result = generate_report(template, csv_base, conn)
            if result:
                generated.append(result)
        except Exception as e:
            print(f"  [ERROR] {e}")
            import traceback
            traceback.print_exc()
    conn.close()
    print(f"\n=== Proceso completado ===")
    print(f"Reportes generados: {len(generated)}")
    for r in generated:
        print(f"  - {r}")


if __name__ == '__main__':
    main()
